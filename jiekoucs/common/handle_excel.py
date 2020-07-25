"""
-- coding: utf-8 --
@Time : 2020/4/13 23:34
@Author : jcoool
@Site : 
@File : handle_excel.py
@Software: PyCharm
"""
import openpyxl


class HandleExcel():
    def __init__(self, filename, sheetname):
        """

        :param filename: 传参：文件路径
        :param sheetname: 传参：表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    # 读取数据
    def read_data(self):
        # 打开文件、表单
        fn = openpyxl.load_workbook(self.filename)
        sn = fn[self.sheetname]
        # 按行读取数据
        rows_data = list(sn.rows)
        # 创建空列表来保存用例数据
        cases_data = []
        # 获取用例标题并存放在列表中
        title_list = []
        for title in rows_data[0]:
            title_list.append(title.value)
        # 获取除表头之外的其他行数据
        for i in rows_data[1:]:
            # 每遍历出来一行数据，就创建一个空列表，来存放该行数据
            data_list = []
            for j in i:
                data_list.append(j.value)
            # 将该行的数据和表头进行打包，转换为字典
            case = dict(zip(title_list, data_list))
            cases_data.append(case)
        return cases_data

    def write_data(self, row, column, value):
        """
        写入数据
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的值
        :return:
        """
        fn = openpyxl.load_workbook(self.filename)
        sn = fn[self.sheetname]
        # 根据行、列去写入内容
        sn.cell(row=row, column=column, value=value)
        # 保存文件
        fn.save(self.filename)

if __name__ == '__main__':
    excel = HandleExcel(r'D:\python\jiekouceshi\jiekoucs\data\apicases.xlsx', 'register')
    data = excel.read_data()
    print(data)
    # a = [1,2,3,4,5,6]
    # print(type(a[0]))
